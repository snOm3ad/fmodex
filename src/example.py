import numpy  as np
import pandas as pd
import matplotlib.pyplot as plt

from sklearn                     import linear_model
from sklearn.metrics             import mean_absolute_error, r2_score, mean_squared_error
from matplotlib                  import font_manager
from openpyxl                    import load_workbook
from pandas                      import Series, DataFrame, DatetimeIndex
from pandas.plotting             import register_matplotlib_converters
from matplotlib.axes             import Axes
from typing                      import Callable, Any
from statsmodels.tsa.ar_model    import AR
from numpy                       import ndarray

register_matplotlib_converters()

def if_any(predictors: DataFrame, unary_predicate: Callable[[DataFrame], bool], irows: bool = True, icols: bool = True):
    # Get the rows for which any of the elements match the predicate
    rows = (unary_predicate(predictors)).any(1)
    # Repeat the same for the columns
    cols = (unary_predicate(predictors)).any(0)
    # Return the data frame that contains the filtered data.
    if   irows and not icols:
        # Return data frame w/ filtered rows only
        return predictors.loc[rows, :]
    elif icols and not irows:
        # Return data frame w/ filtered columns only
        return predictors.loc[:, cols]
    else:
        # Return data frame w/ filtered rows and columns
        return predictors.loc[rows, cols]

def if_all(predictors: DataFrame, unary_predicate: Callable[[DataFrame], bool], irows: bool = True, icols: bool = True):
    # Get the rows for which all of the elements match the predicate
    rows = (unary_predicate(predictors)).all(1)
    # Repeat the same for the columns
    cols = (unary_predicate(predictors)).all(0)
    # Return the data frame that contains the filtered data.
    if   irows and not icols:
        # Return data frame w/ filtered rows only
        return predictors.loc[rows, :]
    elif icols and not irows:
        # Return data frame w/ filtered columns only
        return predictors.loc[:, cols]
    else:
        # Return data frame w/ filtered rows and columns
        return predictors.loc[rows, cols]

def print_header(title: str):
    # The width of the terminal window
    screen_width = 101
    # Just print anything if the title is too big.
    if (len(title) - 2) > screen_width:
        return;
    # Calculate the padding for each side of the title section 
    padding = int((screen_width - len(title) - 2) * 0.5)
    # If the padding is positive the use it.
    print("=" * padding, title, "=" * padding)

def print_details(anything: Any, title: str):
    # First print the header, then print the data
    print_header(title)
    # Print the data and newline
    print(anything, '\n')

def write_to_file(data: DataFrame, sheetname: str):
    # Make this code more easy to write, since we will always write to the same file:
    #   BEFORE: data.loc[outliers.index].to_excel('../data/calculations.xlsx', sheet_name='Outliers')
    #    AFTER: write_to_file(data.loc[outliers.index], sheetname='Outliers')
    #
    # The file where all calculations will be stored.
    master_file = '../data/calculations.xlsx' 
    # Don't open the file if the sheet name is larger than 31 characters.
    if len(sheetname) > 31:
        return;
    # In order to avoid Excel overwritting the file, it must be first opened using
    # the 'load_workbook' function, then passed to the writer object. 
    with pd.ExcelWriter(master_file, engine='openpyxl') as writer:
        writer.book = load_workbook(master_file)
        data.to_excel(writer, sheet_name=sheetname)

def get_results(y_test: ndarray, y_pred: ndarray, model):
    # Calculate the MAE and MSE for the model with current predictor
    MAE, MSE = mean_absolute_error(y_test, y_pred), mean_squared_error(y_test, y_pred)
    # Calculate the 'R^2' score
    R2  = r2_score(y_test, y_pred)
    # Return the list of metrics
    return [model.coef_, model.intercept_, MAE, MSE, R2]

def get_category_matrix(category_vector: Series):
    # Add the categorical variables to the 'X' dataframe
    return pd.get_dummies(category_vector, drop_first=True)

def partition_dates(date_vec: DatetimeIndex, for_training: float):
    # Calculate the half plane
    half_plane = int(for_training * date_vec.size) + 1
    # Check if the data is multidimensional
    # The training dataset goes from the start of the data to the half plane
    data_train = date_vec[:half_plane]
    # The test dataset goes from the half plane to the end.
    data_test  = date_vec[half_plane:]
    return (data_train, data_test)

def partition_data(data, for_training: float):
    # Calculate the half plane
    half_plane = int(for_training * len(data.index)) + 1
    # Check if the data is multidimensional
    if data.ndim < 2:
        # If not, then return the partition of the 1D array reshaped to contain two columns
        # The training dataset goes from the start of the data to the half plane
        data_train = data[:half_plane].to_numpy().reshape(-1, 1)
        # The test dataset goes from the half plane to the end.
        data_test  = data[half_plane:].to_numpy().reshape(-1, 1)
        return (data_train, data_test)
    else:
        # If yes, then return the partition of the N-D array (no reshaping required).
        # The training dataset goes from the start of the data to the half plane
        data_train = data.iloc[:half_plane].to_numpy()
        # The test dataset goes from the half plane to the end.
        data_test  = data.iloc[half_plane:].to_numpy()
        return (data_train, data_test)

def train(for_training: float, Y: Series, X: DataFrame):
    # Split input and output data into training and test data
    y_train, y_test = partition_data(Y, for_training)
    x_train, x_test = partition_data(X, for_training)
    # Create the linear model
    model = linear_model.LinearRegression()
    model.fit(x_train, y_train)
    # Get the predictions using the input test dataset
    y_pred = model.predict(x_test)
    # Get the results 
    return get_results(y_test, y_pred, model)

def train_all(for_training: float, Y: Series, X: DataFrame):
    # Labels for metrics of interest
    metrics = ["Coeffs.", "Intercepts", "MAE", "MSE", "R^2"]
    # Go through every predictor in the input tensor
    predictors = X.columns.to_numpy()
    # Store the resulting metrics in a data frame
    log = DataFrame(index=metrics, columns=predictors)
    # Calculate the regression model for every predictor in the predictors matrix
    for pred_vec in predictors:
        # Write the metrics to the log
        log[pred_vec] = train(for_training, Y, X[pred_vec])
    # Return the dataframe containing the results
    return log

def train_all_categorical(for_training: float, Y: Series, X_descriptive: DataFrame, X_categorical: DataFrame):
    # Labels for metrics of interest
    metrics = ["Coeffs.", "Intercepts", "MAE", "MSE", "R^2"]
    # Go through every predictor in the input tensor
    predictors = X_descriptive.columns.to_numpy()
    # Store the resulting metrics in a data frame for every predictor
    log = DataFrame(index=metrics, columns=predictors)
    # Calculate a regression model for every predictor in the predictors matrix 
    for pred_vec in predictors: 
        # Write the metrics to the log 
        log[pred_vec] = train(for_training, Y, X=pd.concat([X_descriptive[pred_vec], X_categorical], axis=1))
    # Return the dataframe containing the results for all the predictors
    return log 


def simple_correlation_plot(axes: Axes, predictor: Series, acs: Series):
    # Create the title for the plot 
    axes.set_title(predictor.name + " vs Actual Surgeries", fontsize=16, weight='bold', fontname='Roboto')
    # Give the 'x' and 'y' labels for the plot
    axes.set_xlabel(predictor.name, fontsize=12, fontname='Roboto')
    axes.set_ylabel("Actual Surgeries", fontsize=12, fontname='Roboto')
    # Plot the data
    axes.plot(predictor, acs, marker='o', color='r', linestyle='')

def categorical_correlation_plot(axes: Axes, predictor: Series, acs: Series, dow: Series):
    # Create the title for the plot 
    axes.set_title(predictor.name + " vs Actual Surgeries", fontsize=16, weight='bold', fontname='Roboto')
    # Give the 'x' and 'y' labels for the plot
    axes.set_xlabel(predictor.name, fontsize=12, fontname='Roboto')
    axes.set_ylabel('Actual Surgeries', fontsize=12, fontname='Roboto')
    # Group the data by the day of the week
    groups_by_dow = DataFrame({ 'SCS': predictor, 'ACS': acs, 'DOW': dow}).groupby('DOW')
    # Use this for the color map
    colors = {"Mon": "#003542" , "Tue": "#74E0D4", "Wed": "#0000FF", "Thu": "#FA9028", "Fri": "#FF0044"}
    # Plot every group in the groups dataframe
    for name, group in groups_by_dow:
        axes.plot(group['SCS'], group['ACS'], marker='o', c=colors[name], linestyle='', label=name)
    # Show the legend
    axes.legend()

def auto_regression(for_training: float, y_train: Series, date_vector: DatetimeIndex):
    # Create the model
    model = AR(y_train, dates=date_vector)
    fitted_model = model.fit()
    # Make predictions
    y_hat = fitted_model.predict(len(y_train), len(y_train) + len(y_test))
    # Return the predictions
    return y_hat

# In Python, like most programming languages, arrays start at 0
# Start by reading the data from the excel file
raw_data = pd.read_excel('../data/case_data.xlsx')
# Extract the array for the actual number of surgeries from the data frame
actual_surgeries = raw_data['ACTUAL']
# Select the data in range of columns [2, m - 1] where "m = # of columns".
scheduled_surgeries = raw_data.iloc[:, 2:-1]

# Check the deviation of the elements in the predictors:
#       1. Subtract the number of scheduled surgeries from the actual surgeries
#          for every column in the table. 
#       2. Take the absolute value of every element in the table.
error_with_outliers = scheduled_surgeries.sub(actual_surgeries, axis=0).abs()
# Calculate the percentage error (i.e. the percentage of unscheduled surgeries).
percentage_error = error_with_outliers.div(actual_surgeries, axis=0)
# Calculate the mean absolute error for each predictor
MAE = error_with_outliers.mean(axis=0)

# To find ALL the outliers:
#       1. Check the percentage error till 'T-2' days before the surgery
#       2. Identify the rows for which all the elements in them are above 50%
outliers = if_all(percentage_error.loc[:, :'T-2'], lambda x: x > 0.5, icols=False)

# Remove the outliers from the actual surgeries + the scheduled surgeries tables
actual_surgeries = actual_surgeries.drop(outliers.index)
scheduled_surgeries = scheduled_surgeries.drop(outliers.index)
# Extract days data from the raw data array
days_data = raw_data['DOW'].drop(outliers.index)

# Drop the outliers from the error matrix 
error = error_with_outliers.drop(outliers.index)

# Recalculate the percentage error
percentage_error = error.div(actual_surgeries, axis=0)
# Extract the predictors vectors from the scheduled surgeries table.
predictors = if_all(percentage_error, lambda x: x < 0.5, irows=False).columns.to_numpy()

# Recalculate the mean absolute error for each predictor
MAE = error.mean(axis=0) 
# Define the order for the data labels in the 'x' axis.
dow_order = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
fig, axes = plt.subplots(nrows=1, ncols=3, figsize=(16, 6))
# plot the 
categorical_correlation_plot(axes[0], scheduled_surgeries['T-1'], actual_surgeries, raw_data['DOW'].drop(outliers.index))
categorical_correlation_plot(axes[1], scheduled_surgeries['T-2'], actual_surgeries, raw_data['DOW'].drop(outliers.index))
categorical_correlation_plot(axes[2], scheduled_surgeries['T-3'], actual_surgeries, raw_data['DOW'].drop(outliers.index))
plt.show()
